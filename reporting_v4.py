import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

NAVY='17365D'; WHITE='FFFFFF'; BLUE='0000FF'; GREEN='008000'; GRAY='666666'; ORANGE='F4B183'; LIGHTRED='F4CCCC'; TEAL='DDEBF7'; PURPLE='E4DFEC'; LIGHTGREEN='E2F0D9'
MONEY='#,##0;[Red](#,##0);-'; PCT='0.0%;[Red](0.0%);-'; MULT='0.0x;[Red](0.0x);-'


def _section(ws,row,title,end_col=8):
    ws.merge_cells(start_row=row,start_column=1,end_row=row,end_column=end_col)
    c=ws.cell(row,1,title); c.fill=PatternFill('solid',fgColor=NAVY); c.font=Font(color=WHITE,bold=True); c.alignment=Alignment(horizontal='left')


def _hdr(ws,row,n):
    for c in range(1,n+1):
        x=ws.cell(row,c); x.fill=PatternFill('solid',fgColor='D9EAF7'); x.font=Font(bold=True); x.alignment=Alignment(wrap_text=True)


def _setup(ws,widths=None):
    ws.sheet_view.showGridLines=False
    if widths:
        for i,w in enumerate(widths,1): ws.column_dimensions[get_column_letter(i)].width=w


def make_v4_excel(d,nm,adjustments,plan,actions,scenarios=None,diagnostic=None):
    wb=Workbook(); ws=wb.active; ws.title='EXECUTIVE_REPORT'
    _setup(ws,[25,22,22,22,22,22,22,22])
    ws.merge_cells('A1:H1'); ws['A1']='RADAR TURNAROUND 4.1 — REPORT NORMALIZZATO'; ws['A1'].fill=PatternFill('solid',fgColor=NAVY); ws['A1'].font=Font(color=WHITE,bold=True,size=15)
    ws.merge_cells('A3:H4'); ws['A3']='Il report separa dato reported, rettifiche professionali e dato normalizzato. Solo le rettifiche con stato VERIFICATA modificano EBITDA e ricavi operativi. Le proiezioni sono scenari gestionali e non attestazioni.'; ws['A3'].alignment=Alignment(wrap_text=True,vertical='top'); ws['A3'].fill=PatternFill('solid',fgColor=TEAL)
    _section(ws,6,'SNAPSHOT NORMALIZZATO')
    labels=[('Ragione sociale',d.get('ragione_sociale'),'text'),('Ricavi vendite reported',d.get('ricavi_correnti'),'money'),('Valore produzione reported',d.get('valore_produzione'),'money'),('EBITDA reported',nm.get('ebitda_reported'),'money'),('Rettifiche EBITDA verificate',nm.get('rettifiche_ebitda'),'money'),('EBITDA normalizzato',nm.get('ebitda_normalizzato'),'money'),('Ricavi operativi normalizzati',nm.get('ricavi_operativi_normalizzati'),'money'),('EBITDA margin normalizzato',nm.get('ebitda_margin_normalizzato'),'pct'),('CFO reported',d.get('cash_flow_operativo'),'money'),('Debiti tributari',d.get('debiti_tributari'),'money'),('Debiti previdenziali',d.get('debiti_previdenziali'),'money'),('Debiti fornitori',d.get('debiti_fornitori'),'money')]
    r=7
    for lab,val,typ in labels:
        ws.cell(r,1,lab); ws.cell(r,2,val)
        if typ=='money': ws.cell(r,2).number_format=MONEY
        elif typ=='pct': ws.cell(r,2).number_format=PCT
        r+=1
    _section(ws,20,'DIAGNOSI')
    for i,p in enumerate(diagnostic or [],21): ws.merge_cells(start_row=i,start_column=1,end_row=i,end_column=8); ws.cell(i,1,'• '+p); ws.cell(i,1).alignment=Alignment(wrap_text=True)
    _section(ws,27,'TESI OPERATIVA')
    ws.merge_cells('A28:H31'); ws['A28']='La sostenibilità deve essere dimostrata sul core business normalizzato e sulla capacità di generare cassa. Un utile civilistico positivo non sostituisce la verifica di ricorrenza dei ricavi, recuperabilità dei crediti, scadenze del passivo e DSCR prospettico.'; ws['A28'].alignment=Alignment(wrap_text=True,vertical='top')

    raw=wb.create_sheet('DATI_REPORTED'); _setup(raw,[34,20,18,60])
    _section(raw,1,'DATI ESTRATTI / VALIDATI',4); raw.append(['Voce','Valore','Unità','Nota'])
    _hdr(raw,2,4)
    for k,v in d.items():
        raw.append([k,v,'€' if isinstance(v,(int,float)) else '', 'Input reported / validato'])
        if isinstance(v,(int,float)): raw.cell(raw.max_row,2).number_format=MONEY

    adj=wb.create_sheet('RETTIFICHE'); _setup(adj,[18,28,55,18,18,18,18,18,18])
    _section(adj,1,'REGISTRO RETTIFICHE DI NORMALIZZAZIONE',9)
    cols=['Stato','Categoria','Descrizione','Importo','Impatto ricavi','Impatto EBITDA','Ricorrente','Confidenza','Effetto nel modello']
    adj.append(cols); _hdr(adj,2,len(cols))
    for a in adjustments:
        adj.append([a.get('stato'),a.get('categoria'),a.get('descrizione'),a.get('importo'),a.get('impatto_ricavi'),a.get('impatto_ebitda'),a.get('ricorrente'),a.get('confidenza'),'INCLUSO' if str(a.get('stato','')).upper()=='VERIFICATA' else 'ESCLUSO'])
        for c in (4,5,6): adj.cell(adj.max_row,c).number_format=MONEY

    bridge=wb.create_sheet('BRIDGE_NORMALIZZAZIONE'); _setup(bridge,[36,22,60])
    _section(bridge,1,'BRIDGE REPORTED → NORMALIZZATO',3)
    rows=[('EBITDA reported',nm.get('ebitda_reported'),'Base reported'),('Rettifiche EBITDA verificate',nm.get('rettifiche_ebitda'),'Somma delle sole rettifiche VERIFICATE'),('EBITDA normalizzato',nm.get('ebitda_normalizzato'),'Reported + rettifiche'),('Ricavi operativi reported',nm.get('ricavi_operativi_reported'),'Valore della produzione se disponibile'),('Rettifiche ricavi verificate',nm.get('rettifiche_ricavi'),'Somma delle sole rettifiche VERIFICATE'),('Ricavi operativi normalizzati',nm.get('ricavi_operativi_normalizzati'),'Reported + rettifiche'),('EBITDA margin normalizzato',nm.get('ebitda_margin_normalizzato'),'EBITDA norm / ricavi operativi norm')]
    for i,(a,b,c) in enumerate(rows,3): bridge.cell(i,1,a); bridge.cell(i,2,b); bridge.cell(i,3,c); bridge.cell(i,2).number_format=PCT if 'margin' in a.lower() else MONEY

    bp=wb.create_sheet('BUSINESS_PLAN_5Y'); _setup(bp,[10,20,18,18,18,18,18,18,18,18,18,18,18,18])
    _section(bp,1,'BUSINESS PLAN 5 ANNI — SCENARIO BASE',14)
    if plan:
        keys=list(plan[0].keys())
        for c,k in enumerate(keys,1): bp.cell(2,c,k); bp.cell(2,c).font=Font(bold=True); bp.cell(2,c).fill=PatternFill('solid',fgColor='D9EAF7')
        for r,row in enumerate(plan,3):
            for c,k in enumerate(keys,1):
                bp.cell(r,c,row[k])
                if isinstance(row[k],(int,float)):
                    bp.cell(r,c).number_format=PCT if 'margin' in k else (MULT if k=='dscr' else MONEY)

    sc=wb.create_sheet('SCENARI'); _setup(sc,[18,22,22,18,22])
    _section(sc,1,'STRESS TEST',5)
    if scenarios:
        keys=list(scenarios[0].keys())
        for c,k in enumerate(keys,1): sc.cell(2,c,k); sc.cell(2,c).font=Font(bold=True); sc.cell(2,c).fill=PatternFill('solid',fgColor='D9EAF7')
        for r,row in enumerate(scenarios,3):
            for c,k in enumerate(keys,1): sc.cell(r,c,row[k]); sc.cell(r,c).number_format=MULT if 'dscr' in k else MONEY

    act=wb.create_sheet('PIANO_RISANAMENTO'); _setup(act,[24,64,16,24,30])
    _section(act,1,'PIANO ANALITICO DI RISANAMENTO',5)
    act.append(['Area','Azione','Priorità','KPI','Target / trigger']); _hdr(act,2,5)
    for row in actions: act.append(list(row))

    for sht in wb.worksheets:
        sht.freeze_panes='A3'
        for row in sht.iter_rows():
            for cell in row:
                cell.alignment=Alignment(vertical='top',wrap_text=True)
    out=io.BytesIO(); wb.save(out); out.seek(0); return out.getvalue()
