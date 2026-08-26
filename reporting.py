import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

DARK='17365D'; WHITE='FFFFFF'; BLUE='D9EAF7'; ORANGE='FCE4D6'; GREEN='E2F0D9'


def distress_profile(sc):
    if sc.get('composito', 0) < 10:
        return 'NON DISTRESSED'
    vals={'FISCALE':sc.get('fiscale',0)/15,'FINANZIARIA':sc.get('finanziario',0)/15,'COMMERCIALE':sc.get('commerciale',0)/10}
    ordered=sorted(vals.items(), key=lambda x:x[1], reverse=True)
    if ordered[0][1] - ordered[1][1] < 0.12:
        return 'MISTA'
    return ordered[0][0]

def decision_level(d, sc, quality, red_flag=False):
    if red_flag:
        return 'SCARTA / EDD', 'Red flag grave o KO preliminare.'
    if not quality.get('reliable'):
        return 'DATI INSUFFICIENTI', 'Lo scoring non è utilizzabile finché i dati essenziali non sono verificati.'
    rev=d.get('ricavi_correnti'); e=d.get('ebitda'); net=d.get('risultato_netto'); eq=d.get('patrimonio_netto'); ca=d.get('attivo_circolante'); cl=d.get('passivita_correnti')
    # Viability gate: il distress non deve premiare un business strutturalmente distrutto.
    if e is not None and e <= 0 and net is not None and net < 0:
        return 'SCARTA / EDD', 'Viability gate non superata: EBITDA non positivo e perdita netta. Il distress elevato non costituisce di per sé opportunità di turnaround.'
    if rev and eq is not None and eq < -0.25*rev and ca is not None and cl and ca/cl < 0.50:
        return 'SCARTA / EDD', 'Squilibrio patrimoniale e di liquidità estremo: richiede analisi concorsuale/EDD prima di qualsiasi tesi di investimento.'
    # Distress floor: una società sana non deve essere promossa dai punteggi qualitativi neutri.
    if sc.get('composito',0) < 10:
        return 'SCARTA', 'Distress insufficiente rispetto alla strategia Special Situations; non è un target prioritario.'
    if sc['fit']=='FIT-A' and sc['totale'] >= 65 and sc.get('continuita',0) >= 10:
        return 'PRIORITÀ', 'Target coerente con la tesi Tax Turnaround, con continuità economica preliminare sufficiente e meritevole di pre-due-diligence prioritaria.'
    if sc['fit']=='FIT-A' or sc['totale'] >= 60:
        return 'APPROFONDISCI', 'Target da sottoporre a verifica documentale e analisi professionale.'
    if sc['totale'] >= 45:
        return 'MONITORA', 'Profilo non ancora sufficientemente attrattivo; mantenere in watchlist.'
    return 'SCARTA', 'Compatibilità preliminare bassa con il modello di investimento.'


def build_ratios(d):
    rev=d.get('ricavi_correnti'); e=d.get('ebitda'); tax=(d.get('debiti_tributari') or 0)+(d.get('debiti_previdenziali') or 0) if (d.get('debiti_tributari') is not None or d.get('debiti_previdenziali') is not None) else None
    fin=d.get('debito_finanziario'); suppliers=d.get('debiti_fornitori'); ca=d.get('attivo_circolante'); cl=d.get('passivita_correnti')
    return {
        'EBITDA margin': (e/rev) if rev and e is not None else None,
        'Debito fiscale / Ricavi': (tax/rev) if rev and tax is not None else None,
        'Debito fiscale / EBITDA': (tax/e) if e and e>0 and tax is not None else None,
        'Debito finanziario / EBITDA': (fin/e) if e and e>0 and fin is not None else None,
        'Fornitori / Ricavi': (suppliers/rev) if rev and suppliers is not None else None,
        'Current ratio': (ca/cl) if cl and ca is not None else None,
    }


def build_thesis(d, sc):
    r=build_ratios(d)
    strengths=[]; risks=[]
    if d.get('ebitda') is not None and d.get('ebitda')>0: strengths.append('EBITDA positivo: esiste continuità economica residua da verificare e normalizzare.')
    if d.get('ebitda') is not None and d.get('ebitda')<=0: risks.append('EBITDA non positivo: viability gate critica; verificare se esiste un turnaround industriale realistico prima di valorizzare la ristrutturazione del passivo.')
    if r['Debito fiscale / Ricavi'] is not None and r['Debito fiscale / Ricavi']>=0.25: strengths.append('Elevata concentrazione fiscale del passivo, coerente con la specializzazione turnaround fiscale.')
    if sc.get('fit')=='FIT-A': strengths.append('Compatibilità preliminare elevata con la strategia Tax Turnaround.')
    if d.get('patrimonio_netto') is not None and d.get('patrimonio_netto')<0: risks.append('Patrimonio netto negativo: verificare cause, perdite cumulate e presupposti di continuità.')
    if r['Current ratio'] is not None and r['Current ratio']<1: risks.append('Capitale circolante sotto pressione: attivo circolante inferiore alle passività correnti.')
    if r['EBITDA margin'] is not None and r['EBITDA margin']<0.05: risks.append('Marginalità operativa debole: elevata sensibilità a ulteriori shock.')
    if d.get('debiti_fornitori') is not None and d.get('ricavi_correnti') and d['debiti_fornitori']/d['ricavi_correnti']>=0.15: risks.append('Esposizione fornitori significativa rispetto ai ricavi.')
    if not strengths: strengths.append('Nessun elemento positivo sufficiente emerso automaticamente; necessaria valutazione manuale.')
    if not risks: risks.append('Nessuna red flag quantitativa automatica rilevata; restano necessarie verifiche legali, fiscali e industriali.')
    return strengths, risks


def document_checklist(d, sc):
    base=[
        ('Fiscale','Estratto di ruolo / situazione AdER completa','Alta'),
        ('Fiscale','Certificazione debiti Agenzia delle Entrate','Alta'),
        ('Previdenziale','Situazione INPS/INAIL e DURC','Alta'),
        ('Finanziario','Centrale Rischi e dettaglio affidamenti bancari','Alta'),
        ('Commerciale','Ageing fornitori e scaduto per controparte','Alta'),
        ('Commerciale','Ageing clienti e concentrazione primi 10 clienti','Alta'),
        ('Bilancio','Situazione contabile aggiornata infrannuale','Alta'),
        ('Bilancio','Mastri principali e riconciliazione debiti/crediti','Media'),
        ('Legale','Contenzioso civile, tributario, lavoro e amministrativo','Alta'),
        ('Legale','Garanzie reali/personali, fideiussioni e covenant','Alta'),
        ('Corporate','Visura storica, patti, soci, amministratori e parti correlate','Media'),
        ('Operativo','Portafoglio ordini, contratti strategici e autorizzazioni','Media'),
        ('Lavoro','Organico, costo del personale, contenzioso e arretrati','Media'),
        ('Asset','Dettaglio cespiti, magazzino e asset strategici','Media'),
    ]
    if sc.get('fit')=='FIT-A':
        base.insert(2,('Ristrutturazione','Dettaglio per ente, natura, grado e anzianità del debito fiscale/previdenziale','Alta'))
    return base


def make_excel(d, sc, quality, qual, decision, thesis, checklist):
    wb=Workbook(); ws=wb.active; ws.title='SCHEDA_DECISIONALE'
    ws['A1']='RADAR CRISI D’IMPRESA — SCHEDA DECISIONALE'; ws['A1'].font=Font(bold=True,size=16,color=WHITE); ws['A1'].fill=PatternFill('solid',fgColor=DARK); ws.merge_cells('A1:D1')
    rows=[('Ragione sociale',d.get('ragione_sociale')),('Punteggio Radar',sc.get('totale')),('Classe',sc.get('classe')),('Tax Turnaround Fit',sc.get('fit')),('Decisione',decision[0]),('Motivazione',decision[1]),('Qualità dati',f"{quality.get('completeness',0)*100:.0f}%")]
    rr=3
    for k,v in rows:
        ws.cell(rr,1,k).font=Font(bold=True); ws.cell(rr,2,v); rr+=1
    rr+=1; ws.cell(rr,1,'ELEMENTI DI INTERESSE').font=Font(bold=True,color=WHITE); ws.cell(rr,1).fill=PatternFill('solid',fgColor=DARK); ws.merge_cells(start_row=rr,start_column=1,end_row=rr,end_column=4); rr+=1
    for x in thesis[0]: ws.cell(rr,1,'• '+x); ws.merge_cells(start_row=rr,start_column=1,end_row=rr,end_column=4); rr+=1
    rr+=1; ws.cell(rr,1,'CRITICITÀ / RED FLAGS').font=Font(bold=True,color=WHITE); ws.cell(rr,1).fill=PatternFill('solid',fgColor=DARK); ws.merge_cells(start_row=rr,start_column=1,end_row=rr,end_column=4); rr+=1
    for x in thesis[1]: ws.cell(rr,1,'• '+x); ws.merge_cells(start_row=rr,start_column=1,end_row=rr,end_column=4); rr+=1
    for c,w in {'A':35,'B':28,'C':22,'D':22}.items(): ws.column_dimensions[c].width=w

    ws2=wb.create_sheet('DATI_E_SCORING'); ws2.append(['Voce','Valore'])
    for k,v in d.items(): ws2.append([k,v])
    ws2.append([]); ws2.append(['SCORING',''])
    for k,v in sc.items(): ws2.append([k,v])
    ws2.append([]); ws2.append(['QUALITATIVI',''])
    for k,v in qual.items(): ws2.append([k,v])
    ws2.column_dimensions['A'].width=38; ws2.column_dimensions['B'].width=25

    ws3=wb.create_sheet('CHECKLIST_DOCUMENTI'); ws3.append(['Area','Documento / verifica','Priorità','Stato'])
    for row in checklist: ws3.append([row[0],row[1],row[2],'DA RICHIEDERE'])
    for c in range(1,5): ws3.cell(1,c).font=Font(bold=True,color=WHITE); ws3.cell(1,c).fill=PatternFill('solid',fgColor=DARK)
    ws3.column_dimensions['A'].width=22; ws3.column_dimensions['B'].width=72; ws3.column_dimensions['C'].width=14; ws3.column_dimensions['D'].width=18
    out=io.BytesIO(); wb.save(out); out.seek(0); return out
